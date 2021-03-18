from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
wb = load_workbook("sample_formula.xlsx", data_only=True) # 수식이 아닌 데이타 가져온다.
ws = wb.active

# 수식 그대로 가져옴~~~!
for row in ws.values:
  for cell in row:
    print(cell)
