from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Nadosheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# print(ws["A1"]) #A1 셀의 정보 출력(객체)
# print(ws["A1"].value) #A1 셀의 값 출력
# print(ws["A10"].value) # 값 없을때 None

# print(ws.cell(column=1, row=1).value) # ws["A1"] 같다.

# c = ws.cell(column=3, row=1, value=10) # ws["C1"].value =10
# print(c.value)

from random import *

for x in range(1, 11):
  for y in range(1, 11):
    ws.cell(row=x, column=y, value=randint(0, 100))

wb.save("sample.xlsx")