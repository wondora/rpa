from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"]) # 한 줄 데이타 넣기
for i in range(1, 11):
  ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"]
# for cell in col_B:
#   print(cell.value)

# col_range = ws["B:C"]
# for cols in col_range:
#   for cell in cols:
#     print(cell.value)

# row_title = ws[1] # 첫번째 row 만 가져오기
# for cell in row_title:
#   print(cell.value)

# row_range = ws[2:6] # 2부터 6까지 가져오기
# for rows in row_range:
#   for cell in rows:
#     print(cell.value, end=" ")
#   print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#   for cell in rows:
#     # print(cell.value, end=" ")
#     # print(cell.coordinate, end=" ") # 셀의 좌표 정보
#     xy = coordinate_from_string(cell.coordinate) # 듀플로 셀과열을 나눠준다.
#     # print(xy, end=" ") # 셀의 좌표 정보
#     # print("-"*90)
#     print(xy[0], end="") 
#     print(xy[1], end=" ") 
#   print()


#전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#   print(row[1].value)
# tuple(ws.rows)  == ws.iter_rows()

for row in ws.iter_rows(min_row=1, max_row=11, min_col=2, max_col=3):
  print(row[0].value, row[1].value)

#전체 columns
# print(tuple(ws.columns))
# for col in tupe(ws.columns):
#   print(col[1])
# tupe(ws.columns) == ws.iter_cols()


wb.save("sample.xlsx")