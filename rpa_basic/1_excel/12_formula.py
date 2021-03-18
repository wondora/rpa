from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1, 2, 3)"
ws["A3"] = "=AVERAGE(1, 2, 3)"

ws["A4"] = 20
ws["A5"] = 30
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formula.xlsx")