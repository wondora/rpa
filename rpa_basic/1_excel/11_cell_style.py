from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 행의 높이
ws.row_dimensions[1].height =50

a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 취소선
c1.font = Font(color="0000FF", size=20, underline="single") #밑줄 적용

#테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

#90점 넘는 셀에 대해서 초록색
for row in ws.rows:
  for cell in row:
    # 각 셀에 대해 중앙 정렬
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if cell.column == 1: # A 번호열 제외
      continue

    if isinstance(cell.value, int) and cell.value > 80:
      cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
      cell.font = Font(color="FF0000")

#틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정
wb.save("sample_style.xlsx")