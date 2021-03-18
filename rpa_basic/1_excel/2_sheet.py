from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 시트 기본이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # 시트 이름 색깔 변경

#Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")
ws1 = wb.create_sheet("NewSheet", 2) # index 0부터 시작 YourSheet 앞에 생성됨

new_ws = wb["NewSheet"] # Dict 형태로 시트 접근 가능

print(wb.sheetnames)  # 모든 시트 이름 확인

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")